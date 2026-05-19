"""Claude Chat - SSE 流式聊天接口（支持文件上传）"""

import json
import os
import time
import uuid
from pathlib import Path

import httpx
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from app.core.deps import get_current_user

router = APIRouter(prefix="/api/chat", tags=["AI聊天"])

PROXY_URL = "http://127.0.0.1:4000/v1/messages"
MODEL = "claude-sonnet-4-20250514"  # → deepseek-chat via proxy
MAX_TOKENS = 8192
UPLOAD_DIR = Path("/tmp/chat-uploads")
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB

# 支持作为文本读取的文件扩展名
TEXT_EXTENSIONS = {
    ".txt", ".py", ".js", ".ts", ".jsx", ".tsx", ".vue",
    ".css", ".html", ".htm", ".json", ".yaml", ".yml", ".md",
    ".csv", ".xml", ".sql", ".sh", ".bash", ".zsh", ".toml",
    ".ini", ".cfg", ".conf", ".log", ".env", ".envrc",
    ".go", ".rs", ".java", ".c", ".cpp", ".h", ".hpp",
    ".rb", ".php", ".pl", ".lua", ".r", ".m", ".swift",
    ".kt", ".gradle", ".sbt", ".clj", ".ex", ".exs",
    ".proto", ".graphql", ".gql", ".dockerfile",
    ".makefile", ".cmake", ".cmakelists",
}

# 文件大小限制：全文提取上限 100KB（超出截断）
MAX_EXTRACT_SIZE = 100 * 1024


class ChatRequest(BaseModel):
    message: str
    history: list[dict] | None = None
    files: list[dict] | None = None  # [{"file_id": "...", "name": "..."}]


def extract_text_from_file(file_path: Path) -> tuple[str, str]:
    """根据文件类型提取文本，返回 (内容, 文件类型描述)"""
    ext = file_path.suffix.lower()
    size = file_path.stat().st_size

    # ---- 纯文本 / 代码 ----
    if ext in TEXT_EXTENSIONS:
        try:
            text = file_path.read_text("utf-8", errors="replace")
            if len(text) > MAX_EXTRACT_SIZE:
                text = text[:MAX_EXTRACT_SIZE] + f"\n\n... (文件过长，截断至 {MAX_EXTRACT_SIZE//1024}KB)"
            return text, "text"
        except Exception:
            return f"[无法读取文件: {file_path.name}]", "error"

    # ---- PDF ----
    if ext == ".pdf":
        try:
            from pypdf import PdfReader
            reader = PdfReader(str(file_path))
            pages = []
            total = 0
            for page in reader.pages:
                t = page.extract_text() or ""
                if total + len(t) > MAX_EXTRACT_SIZE:
                    t = t[:MAX_EXTRACT_SIZE - total]
                    pages.append(t)
                    pages.append(f"\n\n... (文件过大，提取 {MAX_EXTRACT_SIZE//1024}KB)")
                    break
                pages.append(t)
                total += len(t)
            text = "\n".join(pages)
            if not text.strip():
                text = "[PDF 文件无可提取文本（扫描件/图片型 PDF）]"
            return text, "pdf"
        except Exception as e:
            return f"[PDF 解析失败: {e}]", "error"

    # ---- DOCX ----
    if ext == ".docx":
        try:
            from docx import Document
            doc = Document(str(file_path))
            paras = []
            total = 0
            for p in doc.paragraphs:
                t = p.text.strip()
                if not t:
                    continue
                if total + len(t) > MAX_EXTRACT_SIZE:
                    paras.append(t[:MAX_EXTRACT_SIZE - total])
                    paras.append(f"\n... (文件过大，提取 {MAX_EXTRACT_SIZE//1024}KB)")
                    break
                paras.append(t)
                total += len(t)
            text = "\n".join(paras)
            if not text.strip():
                text = "[DOCX 文件无可提取文本]"
            return text, "docx"
        except Exception as e:
            return f"[DOCX 解析失败: {e}]", "error"

    # ---- XLSX ----
    if ext in (".xlsx", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
            parts = []
            total = 0
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows_text = []
                for row in ws.iter_rows(values_only=True):
                    row_str = " | ".join(str(c) if c is not None else "" for c in row)
                    rows_text.append(row_str)
                sheet_str = f"===== 工作表: {sheet_name} =====\n" + "\n".join(rows_text)
                if total + len(sheet_str) > MAX_EXTRACT_SIZE:
                    sheet_str = sheet_str[:MAX_EXTRACT_SIZE - total]
                    parts.append(sheet_str)
                    parts.append(f"\n... (文件过大，提取 {MAX_EXTRACT_SIZE//1024}KB)")
                    break
                parts.append(sheet_str)
                total += len(sheet_str)
            text = "\n".join(parts)
            if not text.strip():
                text = "[Excel 文件无数据行]"
            return text, "excel"
        except Exception as e:
            return f"[Excel 解析失败: {e}]", "error"

    # ---- 图片（不支持） ----
    if ext in (".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp", ".svg"):
        return f"[图片文件 ({ext})：当前 AI 模型不支持图片识别，仅能根据文件名提供参考]", "image"

    # ---- 二进制/未知 ----
    return f"[不支持的文件格式: {ext}，请上传文本文件、PDF、Word 或 Excel]", "unsupported"


@router.post("/upload")
async def upload_file(
    file: UploadFile = File(...),
    user=Depends(get_current_user),
):
    """上传文件并提取文本内容"""
    if not file.filename:
        raise HTTPException(status_code=400, detail="文件名不能为空")

    # 检查文件大小
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=413, detail=f"文件过大，最大支持 {MAX_FILE_SIZE // 1024 // 1024}MB")

    # 生成唯一 ID
    file_id = uuid.uuid4().hex[:12]
    safe_name = file.filename.replace("..", "").replace("/", "_")
    save_path = UPLOAD_DIR / f"{file_id}_{safe_name}"

    # 保存文件
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    save_path.write_bytes(content)

    # 提取文本
    extracted_text, file_type = extract_text_from_file(save_path)

    # 保存提取结果到元数据文件
    meta = {
        "file_id": file_id,
        "name": safe_name,
        "size": len(content),
        "type": file_type,
        "text": extracted_text,
        "created_at": time.time(),
    }
    meta_path = UPLOAD_DIR / f"{file_id}.meta.json"
    meta_path.write_text(json.dumps(meta, ensure_ascii=False), encoding="utf-8")

    # 返回给前端（不含完整文本，仅元数据）
    return {
        "file_id": file_id,
        "name": safe_name,
        "size": len(content),
        "type": file_type,
        "preview": extracted_text[:200] + "..." if len(extracted_text) > 200 else extracted_text,
    }


def get_file_text(file_id: str, filename: str) -> str | None:
    """根据 file_id 读取之前上传的文本内容"""
    meta_path = UPLOAD_DIR / f"{file_id}.meta.json"
    if not meta_path.exists():
        return None
    try:
        meta = json.loads(meta_path.read_text("utf-8"))
        text = meta.get("text", "")
        name = meta.get("name", filename)
        file_type = meta.get("type", "unknown")
        return f"📄 文件: {name} (类型: {file_type})\n```\n{text}\n```"
    except Exception:
        return None


@router.post("/stream")
async def chat_stream(
    req: ChatRequest,
    user=Depends(get_current_user),
):
    """SSE 流式聊天接口 — 支持文件上传上下文"""

    async def event_stream():
        start_time = time.time()

        # ---- 构建 prompt（含文件内容） ----
        messages = []
        if req.history:
            for msg in req.history:
                messages.append({
                    "role": msg.get("role", "user"),
                    "content": msg.get("content", ""),
                })

        # 如果有文件，构建文件上下文
        user_content = req.message
        if req.files:
            file_context_parts = []
            for f_info in req.files:
                fid = f_info.get("file_id", "")
                fname = f_info.get("name", "未知文件")
                text = get_file_text(fid, fname)
                if text:
                    file_context_parts.append(text)
                else:
                    file_context_parts.append(f"📄 文件: {fname} (文件已过期，请重新上传)")

            if file_context_parts:
                user_content = (
                    "我在以下文件中提到了相关信息，请根据这些文件内容回答我的问题。\n\n"
                    + "========== 上传的文件 ==========\n"
                    + "\n\n".join(file_context_parts)
                    + "\n===============================\n\n"
                    + "我的问题是：\n"
                    + req.message
                )

        messages.append({"role": "user", "content": user_content})

        body = {
            "model": MODEL,
            "messages": messages,
            "stream": True,
            "max_tokens": MAX_TOKENS,
        }

        # 发送开始事件
        yield f"data: {json.dumps({'type': 'start'})}\n\n"

        collected_text = ""
        error_occurred = False

        try:
            async with httpx.AsyncClient(timeout=httpx.Timeout(180.0, connect=10.0)) as client:
                async with client.stream("POST", PROXY_URL, json=body) as resp:
                    if resp.status_code != 200:
                        error_body = await resp.aread()
                        yield f"data: {json.dumps({
                            'type': 'error',
                            'message': f'代理返回错误 ({resp.status_code}): {error_body.decode()}',
                        })}\n\n"
                        return

                    async for line in resp.aiter_lines():
                        current_elapsed = time.time() - start_time
                        if current_elapsed > 175:
                            yield f"data: {json.dumps({
                                'type': 'error',
                                'message': '请求超时（180秒），请简化问题或重试',
                            })}\n\n"
                            error_occurred = True
                            return

                        raw = line.strip()
                        if not raw:
                            continue

                        if raw.startswith("data: "):
                            raw = raw[6:].strip()
                            if not raw:
                                continue

                        try:
                            event = json.loads(raw)
                        except json.JSONDecodeError:
                            continue

                        event_type = event.get("type", "")

                        # ---- 文本增量块 ----
                        if event_type == "content_block_delta":
                            delta = event.get("delta", {})
                            if delta.get("type") == "text_delta":
                                text = delta.get("text", "")
                                if text:
                                    collected_text += text
                                    yield f"data: {json.dumps({
                                        'type': 'text',
                                        'content': text,
                                    })}\n\n"

                        # ---- message_start ----
                        elif event_type == "message_start":
                            msg_data = event.get("message", {})
                            content_blocks = msg_data.get("content", [])
                            for block in content_blocks:
                                if block.get("type") == "text":
                                    text = block.get("text", "")
                                    if text:
                                        chunk = text[len(collected_text):]
                                        if chunk:
                                            collected_text = text
                                            yield f"data: {json.dumps({
                                                'type': 'text',
                                                'content': chunk,
                                            })}\n\n"

                        # ---- message_delta ----
                        elif event_type == "message_delta":
                            usage = event.get("usage", {})
                            elapsed = time.time() - start_time
                            yield f"data: {json.dumps({
                                'type': 'done',
                                'content': collected_text,
                                'tokens_output': usage.get('output_tokens', 0),
                                'duration_ms': int(elapsed * 1000),
                            })}\n\n"

                        # ---- message_stop ----
                        elif event_type == "message_stop":
                            return

        except httpx.ConnectError:
            yield f"data: {json.dumps({
                'type': 'error',
                'message': '无法连接到 AI 代理（127.0.0.1:4000），请确认代理服务是否运行',
            })}\n\n"
            error_occurred = True
        except httpx.TimeoutException:
            yield f"data: {json.dumps({
                'type': 'error',
                'message': 'AI 代理响应超时，请稍后重试',
            })}\n\n"
            error_occurred = True
        except Exception as e:
            yield f"data: {json.dumps({
                'type': 'error',
                'message': f'服务器内部错误: {str(e)}',
            })}\n\n"
            error_occurred = True

        if not error_occurred:
            elapsed = time.time() - start_time
            yield f"data: {json.dumps({
                'type': 'done',
                'content': collected_text,
                'tokens_output': 0,
                'duration_ms': int(elapsed * 1000),
            })}\n\n"

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
            "Content-Type": "text/event-stream; charset=utf-8",
        },
    )
